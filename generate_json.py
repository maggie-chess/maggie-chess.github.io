import json
import re
from uuid import uuid4

from openpyxl import load_workbook

wb = load_workbook(filename='Openings.xlsx')
openings = [x for x in wb.sheetnames if x != 'Template']

result = {x: {} for x in openings}

for opening in openings:
    lines = list(wb[opening].rows)

    for line in lines[1:]:
        variation, pgn = [x.value for x in line]

        if variation is None:
            continue

        if variation not in result[opening].keys():
            result[opening][variation] = [pgn]
        else:
            result[opening][variation].append(pgn)

with open('Openings.json', 'w') as f:
    f.write(json.dumps(result))


# current_html = open('index.html').read()
# with open(current_html, 'w') as f:
#     f.write(re.sub('const openings = .*?\n', f'const openings = {json.dumps(result)}\n'))
